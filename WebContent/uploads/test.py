# from dbconfig import dbcon
# con = dbcon()
# con.connect()

# a = [1,
#      2,
#      3,
#      16,
#      18,
#      20,
#      25,
#      31,
#      33,
#      34,
#      37,
#      40,
#      41,
#      42,
#      43,
#      50,
#      51,
#      52,
#      57,
#      58,
#      59,
#      61,
#      63,
#      64,
#      70,
#      71,
#      72,
#      73,
#      74,
#      78,
#      80,
#      82,
#      84,
#      90,
#      92,
#      93,
#      96,
#      99,
#      100,
#      104,
#      105,
#      106,
#      109,
#      110,
#      117,
#      150,
#      153,
#      154,
#      155,
#      156,
#      158,
#      159,
#      171,
#      181,
#      186,
#      187,
#      188,
#      189,
#      190,
#      191,
#      192,
#      193,
#      196,
#      197,
#      207,
#      209,
#      210,
#      216,
#      217,
#      219,
#      220,
#      221,
#      222,
#      223,
#      224,
#      226,
#      227,
#      229,
#      230,
#      231,
#      232,
#      233,
#      234,
#      235,
#      236,
#      237,
#      238,
#      402,
#      403,
#      1001,
#      1002,
#      1004,
#      1005,
#      1007,
#      1008,
#      1009,
#      1012]

# for user in a:
#     queryToExe = """ insert into xunit.employee (emp_id, name, designation) values (%s,%s,%s) """
#     con.cursor.execute(queryToExe, (user, 'user', 'labour'))
#     con.connection.commit()

# con.disconnect()

from openpyxl import load_workbook
from dbconfig import dbcon
import datetime

wb = load_workbook(filename="december.xlsx")
print(wb.active.title)

sheet = wb.active

ids = []
dates = []
date = datetime.date(2020, 1, 5)
firstdate = 7


for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=2, max_col=2, values_only=True):
    if row[0] is None:
        continue
    else:
        ids.append(row[0])

        for column in sheet.iter_cols(min_row=firstdate, max_row=firstdate, min_col=6, max_col=37, values_only=True):
            if column[0] is None:
                dates.append(0)
            else:
                dates.append(column[0])
            firstdate += 1


print(dates)
print(ids)
